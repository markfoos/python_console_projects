# python3
# SMTP email sender from spreadsheep, in this program it will pull
# requests from an excel file and determine "dues" from the cell,
# then return a "name, date, and email" for each unpaid cell found and send an Email
# for each unpaid cell
# inspired by Automate the Boring Stuff with Python
# Program Created by Mark Foos
# keeping all commented out code to show troubleshooting and logic as this took
# multiple programming attempts
# -----------------------------------------------------------------------------
import openpyxl, smtplib, sys, datetime
# -----------------------------------------------------------------------------

                        # open spreasheet
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
#print(sheet['A1'].value)
                        # define max column and row variables
lastCol = sheet.max_column
#print("The lastcol Value is: " + str(lastCol))
# latestMonth = sheet.cell(row = 1, column = lastCol).value
# print("The latestMonth value is: " + str(latestMonth))
lastRow = sheet.max_row
#print("The lastRow value is: " + str(lastRow))


                        # returns a better looking date for email
def prettyDate(date):
    dateMonth = (date.strftime("%B"))
    dateYear = (date.year)
    formateddDate = (dateMonth + " " + str(dateYear))
    return formateddDate
#print(prettyDate(latestMonth))



#for r in range(2, lastRow + 1):
#    payment = sheet.cell(row = r, column = lastCol).value
#    if payment != 'paid':
#        name = sheet.cell(row = r, column = 1).value
#        email = sheet.cell(row = r, column = 2).value
#
#        unpaidMembers[name] = email
#        print("These are the unpaid dictionary memeber function+")
#        print(unpaidMembers)



                                # open openpyxl function for finding unpaid dues
unpaidMembers = []

for r in sheet.iter_rows(min_row = 2, min_col = 3, max_col = lastCol, max_row = (lastRow)):
    for cell in r:
        if cell.value != 'paid':
#            print("--" * 55)
#            print(cell.value)
#            print(cell.column)
#            print(cell.row)
            column = cell.column
            row = cell.row

            name = sheet.cell(row = row, column = 1).value
#            print(name)
            email = sheet.cell(row = row, column = 2).value
#            print(email)
            date = sheet.cell(row = 1, column = column).value
#            print(date)
            listObject = [name, email, date]
#            print(listObject)

            unpaidMembers.append(listObject)
#            print(unpaidMembers)

            #print(unpaidMembers)

            #name = sheet.cell(row = r, column = 1).value
            #email = sheet.cell(row = r, column = 2).value
            #date = sheet.cell(row = 1, column = sheet.column).value
            #print(date)
            #print(name)


#print(prettyDate(unpaidMembers[0][2]))
#print(".." * 55)


#for x in unpaidMembers:
#    print(x[0])
#    print(x[1])
#    print(x[2])


                                # email variables and function
my_email = input("Enter Email address: ")
my_pw = input("Enter email pw: ")

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(my_email, my_pw)

for x in unpaidMembers:
    name = x[0]
    sendEmail = x[1]
    date = x[2]


    body = """\
    Subject: Hello %s, Payment Noticifation

    Hello %s, There is no record of payment for %s. Please submit at your earliest
    convience!
    Thank you
    """ % (name, name, prettyDate(date))

    print("Sending email to %s..." % sendEmail)
    sendmailStatus = smtpObj.sendmail(my_email, sendEmail, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (sendEmail, sendmailStatus))

smtpObj.quit()



                                # orginal code, outdated documentation


# latestMonth = sheet.cell(row=1, column=lastCol).value

# unpaidMembers = {}

#for r in range(2, sheet.get_highest_row() + 1):
#    payment = sheet.cell(row=r, column=lastCol).value
#    if payment != "paid":
#        name = sheet.cell(row=r, column=1).value
#        email = sheet.cell(row=r, column=1).value
#        unpaidMembers[name] = email

# def getRowMax():
#    openSpreadsheet()
#    highestRow = openSpreadsheet().max_row
#    print(highestRow)
#    return highestRow

#def GetColMax():
#    highestColumn = openSpreadsheet().max_column
#    print(highestColumn)
    # max_columnValue = sheet.cell(row = highestRow, column = highestColumn).value
    # print(max_columnValue)
#    return highestColumn
# first row values
# email = sheet.cell(row = 2, column = 2).value
# print(email)

#for row in sheet.values:
#    for value in row:
    #    if value = "none":
    #        return
#        print(value)

# print(sheet.values)


# c = sheet.cell(row = 1, column = 1)
# print("-" * 50)
# print(c.value)
#def format_first_row(row, cell):
#    for c in row:
#        cell.value = c
#        print(c)

#format_first_row(2, 1)
#for row in sheet.iter_rows(min_row = 1, max_col = 3, max_row = 2, values_only = true):
#    print(row)

#def findCellValue():
    # print("function called")
#    for r in range(1, getRowMax() + 1):
#        # print("intiate")
#        # print(r)
#        for c in range(1, GetColMax() + 1):
#            testCell = openSpreadsheet().cell(row = r, column = c)
#            testValue = testCell.value
#            return testValue, testCell
            #print("this is the testcell value:")
            #print(testCell.value)

#def unpaidValue(cell):
#    if findCellValue() == 1:
#        print("This is testcell")
#        print(cell)
#        print(cell.value)
#        unpaidCell = cell.value
#        return cell




#def main():
    # open WB and sheet function

    # get row max and col max Var function
    # get unpaid cell values

#    dateReturn(unpaidValue(findCellValue()))
    # get name of unpaid cell values

    # get date of unpaid cell values
    # get email of unpaid cell values

#main()

        #for c in range(1, highestColumn + 1):
        #    checkPayment = sheet.cell(row = r, column = c)
        #    print(checkPayment.value)
        #    print("End of checkPayment")

# smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
# smtpObj.ehlo()
# smtpObj.startls()
# smtpObj.login('markdfoos@gmail.com', sys.argv[1])

#for name, email in unpaidMembers.items():
#    body = "Subject %s ues unpaid.Dear %s,Records show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!" %(latestMonth, name, latestMonth)
#    print('Sending email to %s...' % email)
#    sendmailStatus = smtpObj.sendmail('markdfoos@gmail.com', email, body)

#     if sendmailStatus != {}:
#        print('There was a problem sending mail to %s: %s' % (email, sendmailStatus))
# smtpObj.quit()
