#multiplication table application ran in console, returns a new
# workbook with a multiplication table based on the arbv[1] value
#Created by Mark Foos, based on Chapter 12 Assignment in Automate the
# Boring Stuff in Python

import sys, openpyxl
from openpyxl import Workbook


def argvIntVerification(argv):
   try:
       int(argv)
       print("integer is verified")
       return True
   except ValueError:
        return False

for arg in sys.argv[1]:
    if not argvIntVerification(arg):
       sys.exit("All arguements must be integers. Exiting...")



print("opening workbook")


wb = openpyxl.Workbook()

multVar = int(sys.argv[1])

ws = wb.active

sheet = wb.get_sheet_by_name('Sheet')

for i in range(1, (multVar + 1)):
    for j in range(1, (multVar + 1)):
        ws.cell(row = j + 1, column = i + 1, value = (j * i))

#for i in range(1, (multVar + 1)):
#    ws.cell(row = i + 1, column = 1, value = i)

#for rowNum in range(1, multVar + 2):
#    sheet.cell(row = rowNum, column = colNum).value = colNum -1


ws.title = "Multiplication_Table"

# excelFile = input("Name of your Excel File?\n")

wb.save("Multiplication Tables for " + str(multVar) + '.xlsx')


#print(sys.argv[1])


#print(multVar)

#if argv[1] == int:
#    print("integer entered")
#else:
#    print("non integer entered")


#argvIntVerification(multVar)

#print(multVar)
#wb = openpyxl.Workbook()
#sheet = wb.get_active_sheet()
#sheet.title = "Multiplication_Table"
#openpyxl.load_workbook("Multiplication_Table.xlsx")
#wb.save("Multiplication_Table.xlsx")
#ws = wb.active

#ws1 = wb.create_sheet("new_Sheet", 0)

#ws1.title = "Multiplication_Table"
